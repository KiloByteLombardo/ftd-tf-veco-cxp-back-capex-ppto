"""
Módulo de procesamiento de archivos Excel de Prioridades de Pago - Venezuela
Fase 2: Paso 1 (limpiar y devolver) y Paso 2 (montar en template, BigQuery)
"""
import pandas as pd
import numpy as np
import openpyxl
from typing import Optional, Tuple, List, Dict, Any
from io import BytesIO
from pathlib import Path
from datetime import datetime
import os
import json

# Importar funciones de tasa de cambio
from tasa import obtener_tasa_bolivar_dolar, obtener_tasa_euro_dolar, obtener_tasa_peso_colombiano_dolar

# Importar conexión a Google Sheets y BigQuery
from connection import get_google_sheet_data, upload_to_bigquery

# Configuración de carpeta de resultados
RESULTADOS_PATH = Path(__file__).parent.parent / 'resultados'

# Cabezales esperados del archivo Prioridades de Pago
CABEZALES_ESPERADOS = [
    'Numero de Factura',
    'Numero de OC',
    'Tipo Factura',
    'Nombre Lote',
    'Proveedor',
    'RIF',
    'Fecha Documento',
    'Tienda',
    'Sucursal',
    'Monto',
    'Moneda',
    'Fecha Vencimiento',
    'Cuenta',
    'Banco',
    'Id Cta',
    'Método de Pago',
    'Pago Independiente',
    'Prioridad',
    'Monto CAPEX EXT',
    'Monto CAPEX ORD',
    'Monto CADM',
    'Fecha Creación',
    'Solicitante',
    'Proveedor Remito'
]

# ============================================================================
# ARRAYS DE PRIORIDADES PARA CÁLCULOS
# ============================================================================

# Prioridades que mantienen USD como moneda de pago (no se convierten a VES)
PRIORIDADES_USD_MONEDA_PAGO = [69, 70, 71, 72, 73, 74, 75, 76, 77, 87, 86, 88, 84, 85]

# Prioridades que mantienen la cuenta original cuando la moneda es USD
# (incluye 83 adicional para cuenta bancaria)
PRIORIDADES_USD_CUENTA_ORIGINAL = [69, 70, 71, 72, 73, 74, 75, 76, 77, 87, 86, 88, 83, 84, 85]

# Cuenta por defecto cuando USD no está en las prioridades especiales
CUENTA_USD_DEFAULT = "1111"

# Prioridades que mantienen el monto original sin conversión (para Monto Final)
# Si la prioridad está en este array, el monto no se multiplica por la tasa
PRIORIDADES_MONTO_SIN_CONVERSION = [67, 69, 70, 71, 72, 73, 74, 75, 76, 77, 87, 86, 88, 83, 84, 85, 89]

# Margen adicional para tasa de día JUEVES (tasa + 5)
MARGEN_TASA_JUEVES = 5

# ============================================================================
# RENOMBRADO DE COLUMNAS (nombres internos -> nombres de salida en Excel)
# ============================================================================
RENOMBRAR_COLUMNAS = {
    'Tipo Factura': 'Tipo factura',
    'Id Cta': 'CODIGO CTA',
    'Método de Pago': 'METODO DE PAGO',
    'Prioridad': 'Prioridad origen',
    'Monto Capex ORD 2': 'MONTO CAPEX ORD2',
    'Monto Capex EXT 3': 'MONTO CAPEX EXT3',
    'Monto Capex Final': 'MONTO CAPEX FINAL',
    'Monto Opex Final': 'MONTO OPEX FINAL',
    'Dia de Pago': 'Día de pago',
    'Monto Capex ORD USD': 'MONTO CAPEX ORD USD',
    'Monto Capex EXT USD': 'MONTO CAPEX EXT USD',
    'Monto Total USD': 'MONTO TOTAL USD',
    'AREA': 'ÁREA',
    'Tipo Capex': 'TIPO CAPEX',
    'Tipo Capex 2': 'CAPEX',
}

# ============================================================================
# ORDEN DE COLUMNAS PARA EL EXCEL DE SALIDA (PASO 1)
# Usa los nombres NUEVOS (después de renombrar)
# ============================================================================
ORDEN_COLUMNAS_EXCEL = [
    'Numero de Factura',
    'Numero de OC',
    'Tipo factura',
    'Nombre Lote',
    'Proveedor',
    'RIF',
    'Fecha Documento',
    'Tienda',
    'Sucursal',
    'Monto',
    'Moneda',
    'Fecha Vencimiento',
    'Cuenta',
    'CODIGO CTA',
    'METODO DE PAGO',
    'Pago Independiente',
    'Prioridad origen',
    'Monto CAPEX EXT',
    'Monto CAPEX ORD',
    'Monto CADM',
    'Fecha Creación',
    'Solicitante',
    'MONTO CAPEX ORD2',
    'MONTO CAPEX EXT3',
    'MONTO CAPEX FINAL',
    'MONTO OPEX FINAL',
    'Moneda Pago',
    'Monto Final',
    'Cuenta Bancaria',
    'Día de pago',
    'MONTO CAPEX ORD USD',
    'MONTO CAPEX EXT USD',
    'Monto CAPEX USD',
    'Monto OPEX USD',
    'MONTO TOTAL USD',
    'ÁREA',
    'TIPO CAPEX',
    'CAPEX',
]

# ============================================================================
# CONFIGURACIÓN PARA COLUMNA AREA (RECARGAS)
# ============================================================================

# Proveedores que siempre son RECARGAS
PROVEEDORES_SIEMPRE_RECARGAS = [
    "GALAXY ENTERTAINMENT DE VENEZUELA, C.A. (SIMPLE TV )",
    "RECARGAS MOVIL C.A"
]

# Proveedores con condiciones de sucursal para ser RECARGAS
PROVEEDORES_RECARGAS_CONDICIONAL = {
    "CORPORACION DIGITEL, C.A.": ["POSPAGO FACTURA", "PREPAGO RECARGA"],
    "NETUNO, C.A.": ["RECARGAS"],
    "TELEFÓNICA VENEZOLANA, C.A.": ["RECARGAS"],
    "TELEFONICA VENEZOLANA, C.A.": ["RECARGAS"]  # Sin acento por si acaso
}


def renombrar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    """
    Renombra las columnas del DataFrame según RENOMBRAR_COLUMNAS.
    Es idempotente: si las columnas ya tienen los nombres nuevos, no hace nada.
    
    Args:
        df: DataFrame con columnas a renombrar
        
    Returns:
        DataFrame con columnas renombradas
    """
    rename_map = {k: v for k, v in RENOMBRAR_COLUMNAS.items() if k in df.columns}
    if rename_map:
        df = df.rename(columns=rename_map)
        print(f"[PROC] Renombradas {len(rename_map)} columnas")
    return df


def dataframe_a_json_serializable(df: pd.DataFrame) -> List[Dict]:
    """
    Convierte un DataFrame a una lista de diccionarios serializables a JSON.
    Maneja correctamente valores NaT, NaN, Timestamp, etc.
    
    Args:
        df: DataFrame a convertir
        
    Returns:
        Lista de diccionarios serializables
    """
    def convertir_valor(val):
        """Convierte un valor individual a formato serializable."""
        if pd.isna(val):
            return None
        elif isinstance(val, pd.Timestamp):
            return val.isoformat()
        elif isinstance(val, datetime):
            return val.isoformat()
        elif isinstance(val, np.integer):
            return int(val)
        elif isinstance(val, np.floating):
            return float(val) if not np.isnan(val) else None
        elif isinstance(val, np.ndarray):
            return val.tolist()
        else:
            return val
    
    registros = []
    for _, row in df.iterrows():
        registro = {}
        for col in df.columns:
            registro[col] = convertir_valor(row[col])
        registros.append(registro)
    
    return registros


# ============================================================================
# FUNCIONES DE PROCESAMIENTO DE DATAFRAME (Thread 1)
# ============================================================================

def encontrar_cabezales(df_raw: pd.DataFrame, max_filas_busqueda: int = 20) -> Tuple[int, List[str]]:
    """
    Encuentra automáticamente la fila de cabezales iterando por las filas del archivo.
    Busca coincidencias con los cabezales esperados.
    """
    print("[PROC] Buscando cabezales automáticamente...")
    
    for idx in range(min(max_filas_busqueda, len(df_raw))):
        fila = df_raw.iloc[idx]
        valores = [str(v).strip() if pd.notna(v) else '' for v in fila]
        
        # Buscar coincidencias con cabezales esperados
        coincidencias = sum(1 for v in valores if v in CABEZALES_ESPERADOS)
        
        if coincidencias >= 5:  # Al menos 5 cabezales coinciden
            print(f"[PROC] Cabezales encontrados en fila {idx} ({coincidencias} coincidencias)")
            return idx, valores
    
    # Fallback: buscar fila con más valores string no vacíos
    for idx in range(min(max_filas_busqueda, len(df_raw))):
        fila = df_raw.iloc[idx]
        valores_validos = [v for v in fila.dropna() if str(v).strip() != '']
        
        if len(valores_validos) >= 10:
            strings_count = sum(1 for v in valores_validos if isinstance(v, str))
            if strings_count >= len(valores_validos) * 0.5:
                cabezales = [str(v).strip() if pd.notna(v) else f'Columna_{i}' 
                            for i, v in enumerate(fila)]
                print(f"[PROC] Cabezales encontrados en fila {idx} (fallback)")
                return idx, cabezales
    
    print("[PROC] WARN: No se encontraron cabezales, usando fila 0")
    return 0, list(df_raw.columns)


def leer_excel_con_cabezales(file_content: bytes, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Lee un archivo Excel y detecta automáticamente los cabezales.
    """
    print("[PROC] Leyendo archivo Excel...")
    
    df_raw = pd.read_excel(
        BytesIO(file_content),
        sheet_name=sheet_name or 0,
        header=None
    )
    
    print(f"[PROC] Archivo leído: {df_raw.shape[0]} filas x {df_raw.shape[1]} columnas")
    
    header_idx, cabezales = encontrar_cabezales(df_raw)
    
    df = pd.read_excel(
        BytesIO(file_content),
        sheet_name=sheet_name or 0,
        header=header_idx
    )
    
    print(f"[PROC] DataFrame con cabezales: {df.shape[0]} filas x {df.shape[1]} columnas")
    
    return df


def limpiar_datos(df: pd.DataFrame) -> pd.DataFrame:
    """
    Limpia y normaliza los datos del DataFrame.
    """
    print("[PROC] Limpiando datos...")
    
    # Eliminar filas completamente vacías
    df_limpio = df.dropna(how='all')
    filas_eliminadas = len(df) - len(df_limpio)
    
    if filas_eliminadas > 0:
        print(f"[PROC] Eliminadas {filas_eliminadas} filas vacías")
    
    # Eliminar columnas completamente vacías, EXCEPTO las de CABEZALES_ESPERADOS
    cols_vacias = df_limpio.columns[df_limpio.isna().all()]
    cols_a_eliminar = [col for col in cols_vacias if col not in CABEZALES_ESPERADOS]
    if cols_a_eliminar:
        df_limpio = df_limpio.drop(columns=cols_a_eliminar)
        print(f"[PROC] Eliminadas {len(cols_a_eliminar)} columnas vacías no esperadas")
    
    # Limpiar nombres de columnas
    df_limpio.columns = [
        str(col).strip().replace('\n', ' ').replace('\r', '')
        for col in df_limpio.columns
    ]
    
    # Eliminar columnas sin nombre útil
    cols_a_mantener = [col for col in df_limpio.columns if not col.startswith('Unnamed')]
    df_limpio = df_limpio[cols_a_mantener]
    
    # Eliminar filas de resumen (Total de Facturas, Total Facturas, Total) en Numero de Factura
    if 'Numero de Factura' in df_limpio.columns:
        valores_factura = df_limpio['Numero de Factura'].astype(str).str.strip().str.upper()
        mask_total = (
            valores_factura.str.contains('TOTAL DE FACTURAS', na=False) |
            valores_factura.str.contains('TOTAL FACTURAS', na=False) |
            (valores_factura == 'TOTAL')
        )
        filas_total = mask_total.sum()
        if filas_total > 0:
            df_limpio = df_limpio[~mask_total]
            print(f"[PROC] Eliminadas {filas_total} filas de resumen (Total/Total Facturas)")
    
    print(f"[PROC] Datos limpios: {df_limpio.shape[0]} filas x {df_limpio.shape[1]} columnas")
    
    return df_limpio


def calcular_columnas_adicionales(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calcula las 3 columnas adicionales: Moneda Pago, Cuenta Bancaria, Dia de Pago.
    
    Args:
        df: DataFrame con los datos limpios
        
    Returns:
        DataFrame con las columnas adicionales
    """
    print("[PROC] Calculando columnas adicionales...")
    
    df_result = df.copy()
    
    # Obtener columnas necesarias (con manejo de valores nulos)
    moneda = df_result.get('Moneda', pd.Series([''] * len(df_result)))
    prioridad = pd.to_numeric(df_result.get('Prioridad', pd.Series([0] * len(df_result))), errors='coerce').fillna(0).astype(int)
    cuenta = df_result.get('Cuenta', pd.Series([''] * len(df_result)))
    
    # ========================================================================
    # COLUMNA 1: Moneda Pago
    # Lógica: EUR->EUR, COP->COP, USD y prioridad en array->USD, sino->VES
    # ========================================================================
    def calcular_moneda_pago(row_moneda, row_prioridad):
        if pd.isna(row_moneda):
            return 'VES'
        row_moneda = str(row_moneda).strip().upper()
        row_prioridad = int(row_prioridad) if pd.notna(row_prioridad) else 0
        
        if row_moneda == 'EUR':
            return 'EUR'
        elif row_moneda == 'COP':
            return 'COP'
        elif row_moneda == 'USD':
            if row_prioridad in PRIORIDADES_USD_MONEDA_PAGO:
                return 'USD'
            else:
                return 'VES'
        else:
            return 'VES'
    
    df_result['Moneda Pago'] = [
        calcular_moneda_pago(m, p) 
        for m, p in zip(moneda, prioridad)
    ]
    print(f"[PROC] Columna 'Moneda Pago' calculada")
    
    # ========================================================================
    # COLUMNA 2: Cuenta Bancaria
    # Lógica: Si USD y prioridad en array -> cuenta original, si USD -> "1111", sino -> cuenta original
    # ========================================================================
    def calcular_cuenta_bancaria(row_moneda, row_prioridad, row_cuenta):
        if pd.isna(row_moneda):
            return row_cuenta if pd.notna(row_cuenta) else ''
        row_moneda = str(row_moneda).strip().upper()
        row_prioridad = int(row_prioridad) if pd.notna(row_prioridad) else 0
        cuenta_val = str(row_cuenta) if pd.notna(row_cuenta) else ''
        
        if row_moneda == 'USD':
            if row_prioridad in PRIORIDADES_USD_CUENTA_ORIGINAL:
                return cuenta_val
            else:
                return CUENTA_USD_DEFAULT
        else:
            return cuenta_val
    
    df_result['Cuenta Bancaria'] = [
        calcular_cuenta_bancaria(m, p, c) 
        for m, p, c in zip(moneda, prioridad, cuenta)
    ]
    print(f"[PROC] Columna 'Cuenta Bancaria' calculada")
    
    # ========================================================================
    # COLUMNA 3: Dia de Pago
    # Lógica: Si Moneda Pago es USD o EUR -> VIERNES, sino -> JUEVES
    # ========================================================================
    def calcular_dia_pago(moneda_pago):
        if pd.isna(moneda_pago):
            return 'JUEVES'
        moneda_pago = str(moneda_pago).strip().upper()
        
        if moneda_pago in ['USD', 'EUR']:
            return 'VIERNES'
        else:
            return 'JUEVES'
    
    df_result['Dia de Pago'] = [
        calcular_dia_pago(mp) 
        for mp in df_result['Moneda Pago']
    ]
    print(f"[PROC] Columna 'Dia de Pago' calculada")
    
    # ========================================================================
    # OBTENER TASA DE CAMBIO VES/USD
    # ========================================================================
    print("[PROC] Obteniendo tasa de cambio VES/USD...")
    tasa_info = obtener_tasa_bolivar_dolar()
    
    if tasa_info['success'] and tasa_info['tasa']:
        tasa_ves_usd = float(tasa_info['tasa'])
        tasa_ves_usd_mas_5 = tasa_ves_usd + MARGEN_TASA_JUEVES
        print(f"[PROC] Tasa VES/USD: {tasa_ves_usd}, Tasa + 5: {tasa_ves_usd_mas_5}")
    else:
        # Tasa por defecto si falla la consulta
        tasa_ves_usd = 36.50
        tasa_ves_usd_mas_5 = 41.50
        print(f"[PROC] WARN: Usando tasa por defecto: {tasa_ves_usd}")
    
    # Obtener columnas necesarias para Monto Final
    monto = pd.to_numeric(df_result.get('Monto', pd.Series([0] * len(df_result))), errors='coerce').fillna(0)
    capex_ext = pd.to_numeric(df_result.get('Monto CAPEX EXT', pd.Series([0] * len(df_result))), errors='coerce').fillna(0)
    capex_ord = pd.to_numeric(df_result.get('Monto CAPEX ORD', pd.Series([0] * len(df_result))), errors='coerce').fillna(0)
    cadm = pd.to_numeric(df_result.get('Monto CADM', pd.Series([0] * len(df_result))), errors='coerce').fillna(0)
    
    # ========================================================================
    # COLUMNA 4: Monto Final
    # Lógica:
    # - Si Moneda = "VES" → Monto
    # - Si Prioridad está en PRIORIDADES_MONTO_SIN_CONVERSION → Monto
    # - Si Dia de Pago = "JUEVES" → Monto * tasa_ves_usd_mas_5
    # - Sino → Monto * tasa_ves_usd
    # ========================================================================
    def calcular_monto_final(row_moneda, row_prioridad, row_monto, row_dia_pago):
        if pd.isna(row_monto) or row_monto == 0:
            return 0
        
        row_moneda = str(row_moneda).strip().upper() if pd.notna(row_moneda) else ''
        row_prioridad = int(row_prioridad) if pd.notna(row_prioridad) else 0
        row_dia_pago = str(row_dia_pago).strip().upper() if pd.notna(row_dia_pago) else 'JUEVES'
        
        # Si moneda es VES, no hay conversión
        if row_moneda == 'VES':
            return float(row_monto)
        
        # Si prioridad está en el array, no hay conversión
        if row_prioridad in PRIORIDADES_MONTO_SIN_CONVERSION:
            return float(row_monto)
        
        # Aplicar tasa según día de pago
        if row_dia_pago == 'JUEVES':
            return float(row_monto) * tasa_ves_usd_mas_5
        else:
            return float(row_monto) * tasa_ves_usd
    
    df_result['Monto Final'] = [
        calcular_monto_final(m, p, mto, dp) 
        for m, p, mto, dp in zip(moneda, prioridad, monto, df_result['Dia de Pago'])
    ]
    print(f"[PROC] Columna 'Monto Final' calculada")
    
    # Obtener Monto Final como serie para cálculos siguientes
    monto_final = pd.to_numeric(df_result['Monto Final'], errors='coerce').fillna(0)
    
    # ========================================================================
    # COLUMNA 5: Monto Capex Final
    # Lógica:
    # - Si (CAPEX EXT = 0 Y CAPEX ORD = 0) → 0
    # - Sino → ((CAPEX EXT + CAPEX ORD) / (CAPEX EXT + CAPEX ORD + CADM)) * Monto Final
    # ========================================================================
    def calcular_monto_capex_final(row_capex_ext, row_capex_ord, row_cadm, row_monto_final):
        capex_ext_val = float(row_capex_ext) if pd.notna(row_capex_ext) else 0
        capex_ord_val = float(row_capex_ord) if pd.notna(row_capex_ord) else 0
        cadm_val = float(row_cadm) if pd.notna(row_cadm) else 0
        monto_final_val = float(row_monto_final) if pd.notna(row_monto_final) else 0
        
        # Si ambos CAPEX son 0, retornar 0
        if capex_ext_val == 0 and capex_ord_val == 0:
            return 0
        
        # Calcular total
        total = capex_ext_val + capex_ord_val + cadm_val
        
        if total == 0:
            return 0
        
        # Proporción de CAPEX sobre el total
        proporcion_capex = (capex_ext_val + capex_ord_val) / total
        return proporcion_capex * monto_final_val
    
    df_result['Monto Capex Final'] = [
        calcular_monto_capex_final(ce, co, ca, mf) 
        for ce, co, ca, mf in zip(capex_ext, capex_ord, cadm, monto_final)
    ]
    print(f"[PROC] Columna 'Monto Capex Final' calculada")
    
    # ========================================================================
    # COLUMNA 6: Monto Opex Final
    # Lógica:
    # - Si (CAPEX EXT = 0 Y CAPEX ORD = 0) → Monto Final
    # - Sino → (CADM / (CAPEX EXT + CAPEX ORD + CADM)) * Monto Final
    # ========================================================================
    def calcular_monto_opex_final(row_capex_ext, row_capex_ord, row_cadm, row_monto_final):
        capex_ext_val = float(row_capex_ext) if pd.notna(row_capex_ext) else 0
        capex_ord_val = float(row_capex_ord) if pd.notna(row_capex_ord) else 0
        cadm_val = float(row_cadm) if pd.notna(row_cadm) else 0
        monto_final_val = float(row_monto_final) if pd.notna(row_monto_final) else 0
        
        # Si ambos CAPEX son 0, retornar Monto Final completo
        if capex_ext_val == 0 and capex_ord_val == 0:
            return monto_final_val
        
        # Calcular total
        total = capex_ext_val + capex_ord_val + cadm_val
        
        if total == 0:
            return monto_final_val
        
        # Proporción de CADM (OPEX) sobre el total
        proporcion_opex = cadm_val / total
        return proporcion_opex * monto_final_val
    
    df_result['Monto Opex Final'] = [
        calcular_monto_opex_final(ce, co, ca, mf) 
        for ce, co, ca, mf in zip(capex_ext, capex_ord, cadm, monto_final)
    ]
    print(f"[PROC] Columna 'Monto Opex Final' calculada")
    
    # ========================================================================
    # OBTENER TABLA DE ÁREAS DESDE GOOGLE SHEETS
    # ========================================================================
    print("[PROC] Obteniendo tabla de áreas desde Google Sheets...")
    try:
        df_areas = get_google_sheet_data()
        # Crear diccionario para búsqueda rápida (Solicitante -> Area)
        # Asumiendo que columna A es el código y columna B es el área
        areas_dict = {}
        if len(df_areas.columns) >= 2:
            col_codigo = df_areas.columns[0]
            col_area = df_areas.columns[1]
            for _, row in df_areas.iterrows():
                codigo = row[col_codigo]
                area = row[col_area]
                if pd.notna(codigo):
                    areas_dict[str(codigo).strip()] = str(area).strip() if pd.notna(area) else "SERVICIOS"
        print(f"[PROC] Tabla de áreas cargada: {len(areas_dict)} registros")
    except Exception as e:
        print(f"[PROC] WARN: No se pudo cargar tabla de áreas: {str(e)}")
        df_areas = pd.DataFrame()
        areas_dict = {}
    
    # Guardar df_areas en attrs para usarlo en el Excel
    df_result.attrs['df_areas'] = df_areas
    
    # Obtener columnas necesarias para AREA
    proveedor = df_result.get('Proveedor', pd.Series([''] * len(df_result)))
    sucursal = df_result.get('Sucursal', pd.Series([''] * len(df_result)))
    solicitante = df_result.get('Solicitante', pd.Series([''] * len(df_result)))
    
    # Obtener Monto Capex Final y Monto Opex Final para las siguientes columnas
    monto_capex_final = pd.to_numeric(df_result['Monto Capex Final'], errors='coerce').fillna(0)
    monto_opex_final = pd.to_numeric(df_result['Monto Opex Final'], errors='coerce').fillna(0)
    
    # ========================================================================
    # COLUMNA 7: AREA
    # Lógica:
    # - Si proveedor + sucursal cumplen condiciones de RECARGAS → "RECARGAS"
    # - Si Solicitante = 0 → "SERVICIOS"
    # - Si no → BUSCARV en tabla de áreas
    # ========================================================================
    def calcular_area(row_proveedor, row_sucursal, row_solicitante):
        prov = str(row_proveedor).strip().upper() if pd.notna(row_proveedor) else ''
        suc = str(row_sucursal).strip().upper() if pd.notna(row_sucursal) else ''
        sol = str(row_solicitante).strip() if pd.notna(row_solicitante) else '0'
        
        # Verificar proveedores que siempre son RECARGAS
        for prov_recarga in PROVEEDORES_SIEMPRE_RECARGAS:
            if prov_recarga.upper() in prov or prov in prov_recarga.upper():
                return "RECARGAS"
        
        # Verificar proveedores con condiciones de sucursal
        for prov_cond, sucursales in PROVEEDORES_RECARGAS_CONDICIONAL.items():
            if prov_cond.upper() in prov or prov in prov_cond.upper():
                for suc_recarga in sucursales:
                    if suc_recarga.upper() in suc or suc in suc_recarga.upper():
                        return "RECARGAS"
        
        # Si Solicitante es 0 o vacío → SERVICIOS
        try:
            sol_num = float(sol) if sol else 0
            if sol_num == 0:
                return "SERVICIOS"
        except ValueError:
            pass
        
        # Buscar en tabla de áreas
        if sol in areas_dict:
            return areas_dict[sol]
        
        # Si no se encuentra, retornar SERVICIOS
        return "SERVICIOS"
    
    df_result['AREA'] = [
        calcular_area(p, s, sol) 
        for p, s, sol in zip(proveedor, sucursal, solicitante)
    ]
    print(f"[PROC] Columna 'AREA' calculada")
    
    # ========================================================================
    # COLUMNA 8: Tipo Capex 2
    # Lógica:
    # - Si AREA = "RECARGAS" → "RECARGAS"
    # - Si (Monto CAPEX Final <> 0 Y Monto OPEX Final <> 0) → "MIXTA"
    # - Si Monto CAPEX Final <> 0 → "CAPEX"
    # - Si no → "OPEX"
    # ========================================================================
    def calcular_tipo_capex_2(row_area, row_monto_capex_final, row_monto_opex_final):
        area = str(row_area).strip().upper() if pd.notna(row_area) else ''
        mcf = float(row_monto_capex_final) if pd.notna(row_monto_capex_final) else 0
        mof = float(row_monto_opex_final) if pd.notna(row_monto_opex_final) else 0
        
        if area == "RECARGAS":
            return "RECARGAS"
        
        if mcf != 0 and mof != 0:
            return "MIXTA"
        
        if mcf != 0:
            return "CAPEX"
        
        return "OPEX"
    
    df_result['Tipo Capex 2'] = [
        calcular_tipo_capex_2(a, mcf, mof) 
        for a, mcf, mof in zip(df_result['AREA'], monto_capex_final, monto_opex_final)
    ]
    print(f"[PROC] Columna 'Tipo Capex 2' calculada")
    
    # ========================================================================
    # COLUMNA 9: Tipo Capex
    # Lógica:
    # - Si AREA = "RECARGAS" → "RECARGAS"
    # - Si Tipo Capex 2 = "OPEX" → "OPEX"
    # - Si (Monto CAPEX EXT <> 0 Y Monto CAPEX ORD <> 0) → "MIXTA"
    # - Si Monto CAPEX EXT <> 0 → "EXT"
    # - Si no → "ORD"
    # ========================================================================
    def calcular_tipo_capex(row_area, row_tipo_capex_2, row_capex_ext, row_capex_ord):
        area = str(row_area).strip().upper() if pd.notna(row_area) else ''
        tc2 = str(row_tipo_capex_2).strip().upper() if pd.notna(row_tipo_capex_2) else ''
        ce = float(row_capex_ext) if pd.notna(row_capex_ext) else 0
        co = float(row_capex_ord) if pd.notna(row_capex_ord) else 0
        
        if area == "RECARGAS":
            return "RECARGAS"
        
        if tc2 == "OPEX":
            return "OPEX"
        
        if ce != 0 and co != 0:
            return "MIXTA"
        
        if ce != 0:
            return "EXT"
        
        return "ORD"
    
    df_result['Tipo Capex'] = [
        calcular_tipo_capex(a, tc2, ce, co) 
        for a, tc2, ce, co in zip(df_result['AREA'], df_result['Tipo Capex 2'], capex_ext, capex_ord)
    ]
    print(f"[PROC] Columna 'Tipo Capex' calculada")
    
    # ========================================================================
    # COLUMNA 10: Monto Capex ORD 2
    # Lógica:
    # - Si Tipo Capex = "OPEX" o "RECARGAS" o "PRESTAMO" → 0
    # - Si Tipo Capex = "EXT" → 0
    # - Si Tipo Capex = "ORD" → Monto Capex Final
    # - Sino → Monto Capex Final * (CAPEX ORD / (CAPEX ORD + CAPEX EXT))
    # ========================================================================
    def calcular_monto_capex_ord_2(row_tipo_capex, row_monto_capex_final, row_capex_ext, row_capex_ord):
        tc = str(row_tipo_capex).strip().upper() if pd.notna(row_tipo_capex) else ''
        mcf = float(row_monto_capex_final) if pd.notna(row_monto_capex_final) else 0
        ce = float(row_capex_ext) if pd.notna(row_capex_ext) else 0
        co = float(row_capex_ord) if pd.notna(row_capex_ord) else 0
        
        # Si es OPEX, RECARGAS o PRESTAMO → 0
        if tc in ["OPEX", "RECARGAS", "PRESTAMO"]:
            return 0
        
        # Si es EXT → 0
        if tc == "EXT":
            return 0
        
        # Si es ORD → Monto Capex Final
        if tc == "ORD":
            return mcf
        
        # Sino (MIXTA) → Monto Capex Final * (CAPEX ORD / (CAPEX ORD + CAPEX EXT))
        total = co + ce
        if total == 0:
            return 0
        return mcf * (co / total)
    
    df_result['Monto Capex ORD 2'] = [
        calcular_monto_capex_ord_2(tc, mcf, ce, co) 
        for tc, mcf, ce, co in zip(df_result['Tipo Capex'], monto_capex_final, capex_ext, capex_ord)
    ]
    print(f"[PROC] Columna 'Monto Capex ORD 2' calculada")
    
    # ========================================================================
    # COLUMNA 11: Monto Capex EXT 3
    # Lógica:
    # - Si Tipo Capex = "OPEX" o "RECARGAS" o "PRESTAMO" → 0
    # - Si Tipo Capex = "ORD" → 0
    # - Si Tipo Capex = "EXT" → Monto Capex Final
    # - Sino → Monto Capex Final * (CAPEX EXT / (CAPEX ORD + CAPEX EXT))
    # ========================================================================
    def calcular_monto_capex_ext_3(row_tipo_capex, row_monto_capex_final, row_capex_ext, row_capex_ord):
        tc = str(row_tipo_capex).strip().upper() if pd.notna(row_tipo_capex) else ''
        mcf = float(row_monto_capex_final) if pd.notna(row_monto_capex_final) else 0
        ce = float(row_capex_ext) if pd.notna(row_capex_ext) else 0
        co = float(row_capex_ord) if pd.notna(row_capex_ord) else 0
        
        # Si es OPEX, RECARGAS o PRESTAMO → 0
        if tc in ["OPEX", "RECARGAS", "PRESTAMO"]:
            return 0
        
        # Si es ORD → 0
        if tc == "ORD":
            return 0
        
        # Si es EXT → Monto Capex Final
        if tc == "EXT":
            return mcf
        
        # Sino (MIXTA) → Monto Capex Final * (CAPEX EXT / (CAPEX ORD + CAPEX EXT))
        total = co + ce
        if total == 0:
            return 0
        return mcf * (ce / total)
    
    df_result['Monto Capex EXT 3'] = [
        calcular_monto_capex_ext_3(tc, mcf, ce, co) 
        for tc, mcf, ce, co in zip(df_result['Tipo Capex'], monto_capex_final, capex_ext, capex_ord)
    ]
    print(f"[PROC] Columna 'Monto Capex EXT 3' calculada")
    
    # ========================================================================
    # OBTENER TASA EUR/USD
    # ========================================================================
    print("[PROC] Obteniendo tasa EUR/USD...")
    tasa_eur_info = obtener_tasa_euro_dolar()
    if tasa_eur_info['success'] and tasa_eur_info['tasa']:
        tasa_eur_usd = float(tasa_eur_info['tasa'])
        print(f"[PROC] Tasa EUR/USD: {tasa_eur_usd}")
    else:
        tasa_eur_usd = 1.10  # Tasa por defecto
        print(f"[PROC] WARN: Usando tasa EUR/USD por defecto: {tasa_eur_usd}")
    
    # Obtener series necesarias para las nuevas columnas
    monto_capex_ord_2 = pd.to_numeric(df_result['Monto Capex ORD 2'], errors='coerce').fillna(0)
    monto_capex_ext_3 = pd.to_numeric(df_result['Monto Capex EXT 3'], errors='coerce').fillna(0)
    moneda_pago = df_result['Moneda Pago']
    dia_pago = df_result['Dia de Pago']
    
    # ========================================================================
    # COLUMNA 12: Monto Capex ORD USD
    # Lógica:
    # - Si Moneda Pago = "USD" → Monto Capex ORD 2
    # - Si Moneda Pago = "EUR" → Monto Capex ORD 2 * tasa_eur_usd
    # - Si Dia de Pago = "MARTES" → Monto Capex ORD 2 / tasa_ves_usd
    # - Sino → Monto Capex ORD 2 / tasa_ves_usd_mas_5
    # ========================================================================
    def calcular_monto_capex_ord_usd(row_monto, row_moneda_pago, row_dia_pago):
        monto = float(row_monto) if pd.notna(row_monto) else 0
        mp = str(row_moneda_pago).strip().upper() if pd.notna(row_moneda_pago) else ''
        dp = str(row_dia_pago).strip().upper() if pd.notna(row_dia_pago) else ''
        
        if monto == 0:
            return 0
        
        if mp == "USD":
            return monto
        
        if mp == "EUR":
            return monto * tasa_eur_usd
        
        # Para VES u otras monedas, dividir por la tasa
        if dp == "MARTES":
            return monto / tasa_ves_usd if tasa_ves_usd != 0 else 0
        else:
            return monto / tasa_ves_usd_mas_5 if tasa_ves_usd_mas_5 != 0 else 0
    
    df_result['Monto Capex ORD USD'] = [
        calcular_monto_capex_ord_usd(m, mp, dp) 
        for m, mp, dp in zip(monto_capex_ord_2, moneda_pago, dia_pago)
    ]
    print(f"[PROC] Columna 'Monto Capex ORD USD' calculada")
    
    # ========================================================================
    # COLUMNA 13: Monto Capex EXT USD
    # Lógica:
    # - Si Moneda Pago = "USD" → Monto Capex EXT 3
    # - Si Moneda Pago = "COP" → Monto Capex EXT 3 (sin conversión)
    # - Si Moneda Pago = "EUR" → Monto Capex EXT 3 * tasa_eur_usd
    # - Si Dia de Pago = "MARTES" → Monto Capex EXT 3 / tasa_ves_usd
    # - Sino → Monto Capex EXT 3 / tasa_ves_usd_mas_5
    # ========================================================================
    def calcular_monto_capex_ext_usd(row_monto, row_moneda_pago, row_dia_pago):
        monto = float(row_monto) if pd.notna(row_monto) else 0
        mp = str(row_moneda_pago).strip().upper() if pd.notna(row_moneda_pago) else ''
        dp = str(row_dia_pago).strip().upper() if pd.notna(row_dia_pago) else ''
        
        if monto == 0:
            return 0
        
        if mp == "USD":
            return monto
        
        if mp == "COP":
            return monto  # Sin conversión para COP
        
        if mp == "EUR":
            return monto * tasa_eur_usd
        
        # Para VES u otras monedas, dividir por la tasa
        if dp == "MARTES":
            return monto / tasa_ves_usd if tasa_ves_usd != 0 else 0
        else:
            return monto / tasa_ves_usd_mas_5 if tasa_ves_usd_mas_5 != 0 else 0
    
    df_result['Monto Capex EXT USD'] = [
        calcular_monto_capex_ext_usd(m, mp, dp) 
        for m, mp, dp in zip(monto_capex_ext_3, moneda_pago, dia_pago)
    ]
    print(f"[PROC] Columna 'Monto Capex EXT USD' calculada")
    
    # ========================================================================
    # COLUMNA 14: Monto CAPEX USD
    # Lógica:
    # - Si Moneda Pago = "USD" → Monto Capex Final
    # - Si Moneda Pago = "COP" → Monto Capex Final
    # - Si Moneda Pago = "EUR" → Monto Capex Final * tasa_eur_usd
    # - Si Dia de Pago = "MARTES" → Monto Capex Final / tasa_ves_usd
    # - Sino → Monto Capex Final / tasa_ves_usd_mas_5
    # ========================================================================
    def calcular_monto_capex_usd(row_monto, row_moneda_pago, row_dia_pago):
        monto = float(row_monto) if pd.notna(row_monto) else 0
        mp = str(row_moneda_pago).strip().upper() if pd.notna(row_moneda_pago) else ''
        dp = str(row_dia_pago).strip().upper() if pd.notna(row_dia_pago) else ''
        
        if monto == 0:
            return 0
        
        if mp == "USD":
            return monto
        
        if mp == "COP":
            return monto
        
        if mp == "EUR":
            return monto * tasa_eur_usd
        
        # Para VES u otras monedas, dividir por la tasa
        if dp == "MARTES":
            return monto / tasa_ves_usd if tasa_ves_usd != 0 else 0
        else:
            return monto / tasa_ves_usd_mas_5 if tasa_ves_usd_mas_5 != 0 else 0
    
    df_result['Monto CAPEX USD'] = [
        calcular_monto_capex_usd(m, mp, dp) 
        for m, mp, dp in zip(monto_capex_final, moneda_pago, dia_pago)
    ]
    print(f"[PROC] Columna 'Monto CAPEX USD' calculada")
    
    # ========================================================================
    # COLUMNA 15: Monto OPEX USD
    # Lógica:
    # - Si Moneda Pago = "USD" → Monto Opex Final
    # - Si Moneda Pago = "EUR" → Monto Opex Final * tasa_eur_usd
    # - Si Dia de Pago = "MARTES" → Monto Opex Final / tasa_ves_usd
    # - Sino → Monto Opex Final / tasa_ves_usd_mas_5
    # ========================================================================
    def calcular_monto_opex_usd(row_monto, row_moneda_pago, row_dia_pago):
        monto = float(row_monto) if pd.notna(row_monto) else 0
        mp = str(row_moneda_pago).strip().upper() if pd.notna(row_moneda_pago) else ''
        dp = str(row_dia_pago).strip().upper() if pd.notna(row_dia_pago) else ''
        
        if monto == 0:
            return 0
        
        if mp == "USD":
            return monto
        
        if mp == "EUR":
            return monto * tasa_eur_usd
        
        # Para VES u otras monedas, dividir por la tasa
        if dp == "MARTES":
            return monto / tasa_ves_usd if tasa_ves_usd != 0 else 0
        else:
            return monto / tasa_ves_usd_mas_5 if tasa_ves_usd_mas_5 != 0 else 0
    
    df_result['Monto OPEX USD'] = [
        calcular_monto_opex_usd(m, mp, dp) 
        for m, mp, dp in zip(monto_opex_final, moneda_pago, dia_pago)
    ]
    print(f"[PROC] Columna 'Monto OPEX USD' calculada")
    
    # ========================================================================
    # COLUMNA 16: Monto Total USD
    # Lógica: Monto CAPEX USD + Monto OPEX USD
    # ========================================================================
    monto_capex_usd = pd.to_numeric(df_result['Monto CAPEX USD'], errors='coerce').fillna(0)
    monto_opex_usd = pd.to_numeric(df_result['Monto OPEX USD'], errors='coerce').fillna(0)
    
    df_result['Monto Total USD'] = monto_capex_usd + monto_opex_usd
    print(f"[PROC] Columna 'Monto Total USD' calculada")
    
    # ========================================================================
    # OBTENER TASA COP/USD
    # ========================================================================
    print("[PROC] Obteniendo tasa COP/USD...")
    tasa_cop_info = obtener_tasa_peso_colombiano_dolar()
    if tasa_cop_info['success'] and tasa_cop_info['tasa']:
        tasa_cop_usd = float(tasa_cop_info['tasa'])
        print(f"[PROC] Tasa COP/USD: {tasa_cop_usd}")
    else:
        tasa_cop_usd = 0.00024  # Tasa por defecto (1 COP ~ 0.00024 USD)
        print(f"[PROC] WARN: Usando tasa COP/USD por defecto: {tasa_cop_usd}")
    
    # Guardar las tasas en el DataFrame para referencia
    df_result.attrs['tasa_ves_usd'] = tasa_ves_usd
    df_result.attrs['tasa_ves_usd_mas_5'] = tasa_ves_usd_mas_5
    df_result.attrs['tasa_eur_usd'] = tasa_eur_usd
    df_result.attrs['tasa_cop_usd'] = tasa_cop_usd
    
    print(f"[PROC] Columnas adicionales completadas: {df_result.shape[1]} columnas totales")
    
    return df_result


# ============================================================================
# FUNCIONES DE GENERACIÓN DE EXCEL
# ============================================================================

def indice_a_letra_excel(idx: int) -> str:
    """
    Convierte un índice de columna (0-based) a letra de Excel.
    Ej: 0->A, 25->Z, 26->AA, 27->AB
    """
    resultado = ""
    while idx >= 0:
        resultado = chr(65 + (idx % 26)) + resultado
        idx = idx // 26 - 1
    return resultado


def generar_formula_or_prioridades(col_prioridad: str, prioridades: List[int], excel_row: int) -> str:
    """
    Genera la parte OR de la fórmula para verificar múltiples prioridades.
    Ej: OR(T2=69,T2=70,T2=71,...)
    """
    condiciones = [f'{col_prioridad}{excel_row}={p}' for p in prioridades]
    return f'OR({",".join(condiciones)})'


def crear_excel_con_formulas(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Crea un archivo Excel en memoria con la hoja 'Detalle' y datos calculados.
    El DataFrame ya incluye todas las columnas calculadas (Monto Capex Final, Monto Opex Final, etc.)
    
    Returns:
        Dict con excel_bytes y metadata del archivo generado
    """
    print("[EXCEL] Creando Excel con datos...")
    
    # Renombrar columnas según el mapeo
    df = renombrar_columnas(df)
    
    # Reordenar columnas según el orden deseado (y excluir Banco, Proveedor Remito)
    columnas_disponibles = [col for col in ORDEN_COLUMNAS_EXCEL if col in df.columns]
    df = df[columnas_disponibles]
    
    # Crear el archivo Excel en memoria con xlsxwriter
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Escribir datos en hoja 'Detalle' (el DataFrame ya incluye todas las columnas calculadas)
        df.to_excel(writer, sheet_name='Detalle', index=False, startrow=0)
        
        workbook = writer.book
        worksheet = writer.sheets['Detalle']
        
        # Formatos
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        money_format = workbook.add_format({
            'num_format': '#,##0.00',
            'border': 1
        })
        
        text_format = workbook.add_format({
            'border': 1
        })
        
        formula_format = workbook.add_format({
            'bold': True,
            'bg_color': '#E2EFDA',
            'num_format': '#,##0.00',
            'border': 1
        })
        
        # Aplicar formato a cabezales
        for col_num, col_name in enumerate(df.columns):
            worksheet.write(0, col_num, col_name, header_format)
        
        # Ajustar ancho de columnas
        for col_num, col_name in enumerate(df.columns):
            max_length = max(len(str(col_name)), 12)
            worksheet.set_column(col_num, col_num, max_length + 2)
        
        # Variables para el Excel
        num_filas = len(df)
        num_cols = len(df.columns)
        
        # Encontrar índices de columnas (tanto originales como calculadas)
        col_indices = {col: idx for idx, col in enumerate(df.columns)}
        
        # Columnas calculadas (ya en el DataFrame, con nombres nuevos)
        moneda_pago_col = col_indices.get('Moneda Pago')
        cuenta_bancaria_col = col_indices.get('Cuenta Bancaria')
        dia_pago_col = col_indices.get('Día de pago')
        monto_final_col = col_indices.get('Monto Final')
        monto_capex_final_col = col_indices.get('MONTO CAPEX FINAL')
        monto_opex_final_col = col_indices.get('MONTO OPEX FINAL')
        area_col = col_indices.get('ÁREA')
        tipo_capex_2_col = col_indices.get('CAPEX')
        tipo_capex_col = col_indices.get('TIPO CAPEX')
        monto_capex_ord_2_col = col_indices.get('MONTO CAPEX ORD2')
        monto_capex_ext_3_col = col_indices.get('MONTO CAPEX EXT3')
        monto_capex_ord_usd_col = col_indices.get('MONTO CAPEX ORD USD')
        monto_capex_ext_usd_col = col_indices.get('MONTO CAPEX EXT USD')
        monto_capex_usd_col = col_indices.get('Monto CAPEX USD')
        monto_opex_usd_col = col_indices.get('Monto OPEX USD')
        monto_total_usd_col = col_indices.get('MONTO TOTAL USD')
        
        # Aplicar formato de moneda a las columnas numéricas calculadas
        columnas_moneda = [
            monto_final_col, monto_capex_final_col, monto_opex_final_col,
            monto_capex_ord_2_col, monto_capex_ext_3_col,
            monto_capex_ord_usd_col, monto_capex_ext_usd_col,
            monto_capex_usd_col, monto_opex_usd_col, monto_total_usd_col
        ]
        
        for col_idx in columnas_moneda:
            if col_idx is not None:
                worksheet.set_column(col_idx, col_idx, 18, money_format)
        
        print(f"[EXCEL] Columnas del DataFrame escritas: {num_cols} columnas")
        
        # ====================================================================
        # CREAR HOJA "Tasa" CON LAS TASAS DE CAMBIO
        # ====================================================================
        # Obtener tasas de cambio VES/USD
        tasa_info = obtener_tasa_bolivar_dolar()
        if tasa_info['success'] and tasa_info['tasa']:
            tasa_ves_usd = float(tasa_info['tasa'])
        else:
            tasa_ves_usd = 50.0  # Tasa por defecto si falla la API
        tasa_ves_usd_mas_5 = tasa_ves_usd + MARGEN_TASA_JUEVES
        
        # Obtener tasa EUR/USD
        tasa_eur_info = obtener_tasa_euro_dolar()
        if tasa_eur_info['success'] and tasa_eur_info['tasa']:
            tasa_eur_usd = float(tasa_eur_info['tasa'])
        else:
            tasa_eur_usd = 1.10  # Tasa por defecto
        
        print(f"[EXCEL] Tasas obtenidas: VES/USD={tasa_ves_usd}, VES+5={tasa_ves_usd_mas_5}, EUR/USD={tasa_eur_usd}")
        
        # Crear hoja "Tasa"
        ws_tasa = workbook.add_worksheet('Tasa')
        
        # Formatos para la hoja Tasa
        tasa_header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_size': 12
        })
        
        tasa_value_format = workbook.add_format({
            'num_format': '#,##0.00',
            'border': 1,
            'align': 'center',
            'font_size': 14,
            'bold': True
        })
        
        tasa_label_format = workbook.add_format({
            'border': 1,
            'align': 'left',
            'font_size': 11
        })
        
        # Escribir encabezados y valores en hoja Tasa
        ws_tasa.write(0, 0, 'Descripción', tasa_header_format)
        ws_tasa.write(0, 1, 'Valor', tasa_header_format)
        ws_tasa.write(0, 2, 'Fuente', tasa_header_format)
        
        # Fila 2: Tasa VES/USD (celda B2)
        ws_tasa.write(1, 0, 'Tasa VES/USD', tasa_label_format)
        ws_tasa.write(1, 1, tasa_ves_usd, tasa_value_format)
        ws_tasa.write(1, 2, tasa_info.get('fuente', 'DolarAPI') if tasa_info['success'] else 'Por defecto', tasa_label_format)
        
        # Fila 3: Tasa VES/USD + 5 (celda B3)
        ws_tasa.write(2, 0, 'Tasa VES/USD + 5', tasa_label_format)
        ws_tasa.write(2, 1, tasa_ves_usd_mas_5, tasa_value_format)
        ws_tasa.write(2, 2, 'Calculada (Tasa + 5)', tasa_label_format)
        
        # Fila 4: Margen
        ws_tasa.write(3, 0, 'Margen día JUEVES', tasa_label_format)
        ws_tasa.write(3, 1, MARGEN_TASA_JUEVES, tasa_value_format)
        ws_tasa.write(3, 2, 'Configuración', tasa_label_format)
        
        # Fila 5: Tasa EUR/USD (celda B5)
        ws_tasa.write(4, 0, 'Tasa EUR/USD', tasa_label_format)
        ws_tasa.write(4, 1, tasa_eur_usd, tasa_value_format)
        ws_tasa.write(4, 2, tasa_eur_info.get('fuente', 'Frankfurter') if tasa_eur_info['success'] else 'Por defecto', tasa_label_format)
        
        # Fila 6: Fecha de consulta
        ws_tasa.write(5, 0, 'Fecha consulta', tasa_label_format)
        ws_tasa.write(5, 1, tasa_info.get('fecha', datetime.now().strftime('%Y-%m-%d')), tasa_label_format)
        ws_tasa.write(5, 2, tasa_info.get('timestamp', ''), tasa_label_format)
        
        # Ajustar ancho de columnas
        ws_tasa.set_column(0, 0, 20)
        ws_tasa.set_column(1, 1, 15)
        ws_tasa.set_column(2, 2, 30)
        
        print(f"[EXCEL] Hoja 'Tasa' creada")
        
        # ====================================================================
        # CREAR HOJA "Areas" CON LA TABLA DE ÁREAS
        # ====================================================================
        print("[EXCEL] Obteniendo tabla de áreas desde Google Sheets...")
        try:
            df_areas = get_google_sheet_data()
            
            # Crear hoja "Areas"
            ws_areas = workbook.add_worksheet('Areas')
            
            # Escribir encabezados
            for col_idx, col_name in enumerate(df_areas.columns):
                ws_areas.write(0, col_idx, col_name, header_format)
            
            # Escribir datos
            for row_idx, row in df_areas.iterrows():
                for col_idx, value in enumerate(row):
                    ws_areas.write(row_idx + 1, col_idx, value, text_format)
            
            # Ajustar ancho de columnas
            for col_idx in range(len(df_areas.columns)):
                ws_areas.set_column(col_idx, col_idx, 20)
            
            num_areas = len(df_areas)
            print(f"[EXCEL] Hoja 'Areas' creada con {num_areas} registros")
        except Exception as e:
            print(f"[EXCEL] WARN: No se pudo crear hoja 'Areas': {str(e)}")
            df_areas = pd.DataFrame()
            num_areas = 0
        
        # Freeze panes (fijar encabezado)
        worksheet.freeze_panes(1, 0)
        
        print(f"[EXCEL] Excel creado con {num_filas} filas y {num_cols} columnas")
    
    output.seek(0)
    
    return {
        'excel_bytes': output.getvalue(),
        'filas': num_filas,
        'columnas': num_cols,
        'tasas': {
            'tasa_ves_usd': tasa_ves_usd,
            'tasa_ves_usd_mas_5': tasa_ves_usd_mas_5
        },
        'columnas_calculadas': [
            'Moneda Pago',
            'Cuenta Bancaria', 
            'Día de pago',
            'Monto Final',
            'MONTO CAPEX FINAL',
            'MONTO OPEX FINAL',
            'ÁREA',
            'CAPEX',
            'TIPO CAPEX',
            'MONTO CAPEX ORD2',
            'MONTO CAPEX EXT3',
            'MONTO CAPEX ORD USD',
            'MONTO CAPEX EXT USD',
            'Monto CAPEX USD',
            'Monto OPEX USD',
            'MONTO TOTAL USD'
        ],
        'hojas_adicionales': ['Tasa', 'Areas']
    }


# ============================================================================
# FUNCIÓN: MONTAR DATOS EN TEMPLATE (Paso 2)
# ============================================================================

def montar_data_en_template(df: pd.DataFrame, template_bytes: bytes, tasas: Dict[str, float] = None) -> bytes:
    """
    Monta los datos del DataFrame procesado en la hoja 'Detalle' de un template Excel.
    Usa openpyxl para abrir el template existente y escribir los datos.
    Tambien escribe las tasas de cambio en celdas especificas de la hoja 'Detalle'.
    
    Args:
        df: DataFrame procesado con los datos a montar
        template_bytes: bytes del template Excel descargado de GCS
        tasas: Diccionario con las tasas de cambio:
            - tasa_ves_usd: Tasa martes (VES/USD)
            - tasa_ves_usd_mas_5: Tasa jueves (VES/USD + 5)
            - tasa_eur_usd: Tasa EUR/USD
            - tasa_cop_usd: Tasa COP/USD
        
    Returns:
        bytes del Excel final con los datos montados en 'Detalle'
    """
    print("[PASO2] Montando datos en template...")
    
    # Abrir template con openpyxl
    wb = openpyxl.load_workbook(BytesIO(template_bytes))
    
    # Acceder a la hoja 'Detalle'
    if 'Detalle' not in wb.sheetnames:
        raise ValueError(
            f"La hoja 'Detalle' no existe en el template. "
            f"Hojas disponibles: {wb.sheetnames}"
        )
    
    ws = wb['Detalle']
    
    # NO escribir cabezales: el template ya tiene sus propios cabezales.
    # Solo escribir datos a partir de la columna D (columna 4) y fila 2.
    COL_INICIO = 4  # Columna D
    FILA_INICIO = 2  # Fila 2 (fila 1 son cabezales del template)
    
    for row_idx, (_, row) in enumerate(df.iterrows(), start=FILA_INICIO):
        for col_offset, col_name in enumerate(df.columns):
            value = row[col_name]
            col_destino = COL_INICIO + col_offset  # D=4, E=5, F=6, ...
            # Convertir valores numpy/pandas a tipos nativos de Python
            if pd.isna(value):
                ws.cell(row=row_idx, column=col_destino, value=None)
            elif isinstance(value, (np.integer,)):
                ws.cell(row=row_idx, column=col_destino, value=int(value))
            elif isinstance(value, (np.floating,)):
                ws.cell(row=row_idx, column=col_destino, value=float(value))
            elif isinstance(value, pd.Timestamp):
                ws.cell(row=row_idx, column=col_destino, value=value.to_pydatetime())
            elif isinstance(value, datetime):
                ws.cell(row=row_idx, column=col_destino, value=value)
            else:
                ws.cell(row=row_idx, column=col_destino, value=value)
    
    print(f"[PASO2] Datos montados: {len(df)} filas x {len(df.columns)} columnas desde columna D en hoja 'Detalle'")
    
    # ========================================================================
    # ESCRIBIR TASAS DE CAMBIO EN CELDAS ESPECÍFICAS DE LA HOJA 'Detalle'
    # AQ1 = Tasa martes (VES/USD)
    # AT1 = Tasa jueves (VES/USD + 5)
    # AW1 = Tasa EUR/USD
    # AZ1 = Tasa COP/USD
    # ========================================================================
    if tasas:
        tasa_martes = tasas.get('tasa_ves_usd', 0)
        tasa_jueves = tasas.get('tasa_ves_usd_mas_5', 0)
        tasa_eur = tasas.get('tasa_eur_usd', 0)
        tasa_cop = tasas.get('tasa_cop_usd', 0)
        
        ws['AQ1'] = tasa_martes
        ws['AT1'] = tasa_jueves
        ws['AW1'] = tasa_eur
        ws['AZ1'] = tasa_cop
        
        print(f"[PASO2] Tasas escritas en Detalle: AQ1={tasa_martes}, AT1={tasa_jueves}, AW1={tasa_eur}, AZ1={tasa_cop}")
    else:
        print("[PASO2] WARN: No se proporcionaron tasas para escribir en el template")
    
    # Guardar a BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    
    return output.getvalue()


# ============================================================================
# PASO 1: LIMPIAR Y DEVOLVER
# ============================================================================

def procesar_paso1(file_content: bytes, sheet_name: Optional[str] = None) -> dict:
    """
    Paso 1: Procesa/limpia el archivo Excel y lo devuelve.
    NO sube a BigQuery. El archivo procesado se guarda en GCS /tmp desde api.py.
    
    Args:
        file_content: Contenido del archivo Excel en bytes
        sheet_name: Nombre de la hoja (opcional)
        
    Returns:
        Diccionario con:
        - success: bool
        - excel_bytes: bytes del Excel procesado
        - excel_info: metadata del Excel generado
        - stats: estadísticas del procesamiento
        - data: datos en formato JSON serializable
    """
    print("=" * 70)
    print("[PASO1] Iniciando procesamiento de Prioridades de Pago - Venezuela")
    print("[PASO1] Modo: Limpiar y devolver (sin BigQuery)")
    print("=" * 70)
    
    try:
        # 1. Leer Excel con cabezales
        df = leer_excel_con_cabezales(file_content, sheet_name)
        
        # 2. Limpiar datos
        df_limpio = limpiar_datos(df)
        
        # 3. Calcular columnas adicionales
        df_procesado = calcular_columnas_adicionales(df_limpio)
        
        # 4. Generar Excel procesado (en memoria)
        excel_result = crear_excel_con_formulas(df_procesado)
        
        # 5. Calcular estadísticas
        stats = {
            'total_filas': len(df_procesado),
            'total_columnas': len(df_procesado.columns),
            'columnas': list(df_procesado.columns),
            'montos': {},
            'resumen_moneda_pago': df_procesado['Moneda Pago'].value_counts().to_dict() if 'Moneda Pago' in df_procesado.columns else {},
            'resumen_dia_pago': df_procesado['Dia de Pago'].value_counts().to_dict() if 'Dia de Pago' in df_procesado.columns else {},
            'tasas': {
                'tasa_ves_usd': df_procesado.attrs.get('tasa_ves_usd', 0),
                'tasa_ves_usd_mas_5': df_procesado.attrs.get('tasa_ves_usd_mas_5', 0),
                'tasa_eur_usd': df_procesado.attrs.get('tasa_eur_usd', 0),
                'tasa_cop_usd': df_procesado.attrs.get('tasa_cop_usd', 0),
            }
        }
        
        # Calcular sumas de montos
        columnas_monto = ['Monto', 'Monto CAPEX EXT', 'Monto CAPEX ORD', 'Monto CADM']
        for col in columnas_monto:
            if col in df_procesado.columns:
                try:
                    stats['montos'][col] = float(pd.to_numeric(df_procesado[col], errors='coerce').sum())
                except Exception:
                    stats['montos'][col] = 0
        
        print(f"[PASO1] Procesamiento completado: {stats['total_filas']} filas")
        
        return {
            'success': True,
            'message': 'Archivo procesado correctamente (Paso 1)',
            'excel_bytes': excel_result['excel_bytes'],
            'excel_info': {
                'filas': excel_result['filas'],
                'columnas': excel_result['columnas'],
                'tasas': excel_result['tasas'],
                'columnas_calculadas': excel_result['columnas_calculadas'],
                'hojas_adicionales': excel_result['hojas_adicionales'],
            },
            'stats': stats,
            'data': dataframe_a_json_serializable(df_procesado)
        }
        
    except Exception as e:
        print(f"[PASO1] ERROR: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'excel_bytes': None,
            'data': None
        }


# ============================================================================
# PASO 2: MONTAR EN TEMPLATE Y PREPARAR PARA BIGQUERY
# ============================================================================

def procesar_paso2(file_content: bytes, template_bytes: bytes, sheet_name: Optional[str] = None) -> dict:
    """
    Paso 2: Recibe el archivo procesado (del paso 1), lo monta en el template
    de GCS y prepara el DataFrame para subir a BigQuery.
    
    Args:
        file_content: Contenido del archivo Excel procesado (del paso 1, posiblemente editado)
        template_bytes: bytes del template descargado de GCS
        sheet_name: Nombre de la hoja del archivo procesado (opcional)
        
    Returns:
        Diccionario con:
        - success: bool
        - excel_bytes: bytes del Excel final con template
        - df: DataFrame para subir a BigQuery
        - stats: estadísticas
    """
    print("=" * 70)
    print("[PASO2] Iniciando montaje en template y preparación para BigQuery")
    print("=" * 70)
    
    try:
        # 1. Leer el archivo procesado
        df = leer_excel_con_cabezales(file_content, sheet_name)
        df_limpio = limpiar_datos(df)
        
        # Verificar si el archivo ya tiene las columnas calculadas (del paso 1)
        # Usa 'ÁREA' (nombre nuevo después de renombrar en paso 1)
        columnas_calculadas = ['Moneda Pago', 'Monto Final', 'ÁREA']
        tiene_columnas = all(col in df_limpio.columns for col in columnas_calculadas)
        
        if tiene_columnas:
            print("[PASO2] Archivo ya contiene columnas calculadas, usando datos existentes")
            df_procesado = df_limpio
            # Obtener tasas consultando las APIs (ya que attrs no se persisten en el Excel)
            print("[PASO2] Obteniendo tasas de cambio para el template...")
            from tasa import obtener_tasa_bolivar_dolar, obtener_tasa_euro_dolar, obtener_tasa_peso_colombiano_dolar
            
            tasa_info = obtener_tasa_bolivar_dolar()
            tasa_ves_usd = float(tasa_info['tasa']) if tasa_info['success'] and tasa_info['tasa'] else 36.50
            tasa_ves_usd_mas_5 = tasa_ves_usd + MARGEN_TASA_JUEVES
            
            tasa_eur_info = obtener_tasa_euro_dolar()
            tasa_eur_usd = float(tasa_eur_info['tasa']) if tasa_eur_info['success'] and tasa_eur_info['tasa'] else 1.10
            
            tasa_cop_info = obtener_tasa_peso_colombiano_dolar()
            tasa_cop_usd = float(tasa_cop_info['tasa']) if tasa_cop_info['success'] and tasa_cop_info['tasa'] else 0.00024
        else:
            print("[PASO2] Recalculando columnas adicionales...")
            df_procesado = calcular_columnas_adicionales(df_limpio)
            df_procesado = renombrar_columnas(df_procesado)
            # Obtener tasas desde los attrs del DataFrame (fueron guardadas por calcular_columnas_adicionales)
            tasa_ves_usd = df_procesado.attrs.get('tasa_ves_usd', 0)
            tasa_ves_usd_mas_5 = df_procesado.attrs.get('tasa_ves_usd_mas_5', 0)
            tasa_eur_usd = df_procesado.attrs.get('tasa_eur_usd', 0)
            tasa_cop_usd = df_procesado.attrs.get('tasa_cop_usd', 0)
        
        # Preparar diccionario de tasas para el template
        tasas = {
            'tasa_ves_usd': tasa_ves_usd,
            'tasa_ves_usd_mas_5': tasa_ves_usd_mas_5,
            'tasa_eur_usd': tasa_eur_usd,
            'tasa_cop_usd': tasa_cop_usd,
        }
        
        # Guardar tasas en attrs del DataFrame para BigQuery
        df_procesado.attrs['tasa_ves_usd'] = tasa_ves_usd
        df_procesado.attrs['tasa_ves_usd_mas_5'] = tasa_ves_usd_mas_5
        df_procesado.attrs['tasa_eur_usd'] = tasa_eur_usd
        df_procesado.attrs['tasa_cop_usd'] = tasa_cop_usd
        
        # 2. Montar datos en el template (incluyendo tasas en celdas AQ1, AT1, AW1, AZ1)
        excel_bytes = montar_data_en_template(df_procesado, template_bytes, tasas)
        
        # 3. Estadísticas
        stats = {
            'total_filas': len(df_procesado),
            'total_columnas': len(df_procesado.columns),
            'columnas': list(df_procesado.columns),
        }
        
        print(f"[PASO2] Procesamiento completado: {stats['total_filas']} filas montadas en template")
        
        return {
            'success': True,
            'message': 'Datos montados en template correctamente (Paso 2)',
            'excel_bytes': excel_bytes,
            'df': df_procesado,
            'stats': stats
        }
        
    except Exception as e:
        print(f"[PASO2] ERROR: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'excel_bytes': None,
            'df': None,
            'stats': None
        }


# ============================================================================
# FUNCIÓN AUXILIAR: OBTENER SOLO EL DATAFRAME
# ============================================================================

def obtener_dataframe(file_content: bytes, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Obtiene el DataFrame procesado y limpio, listo para usar.
    
    Args:
        file_content: Contenido del archivo Excel en bytes
        sheet_name: Nombre de la hoja (opcional)
        
    Returns:
        DataFrame procesado
    """
    df = leer_excel_con_cabezales(file_content, sheet_name)
    df_limpio = limpiar_datos(df)
    return df_limpio
